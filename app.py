import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# 1. Enter the spreadsheet and extract the client's CPF.
client_sheet = openpyxl.load_workbook('dados_clientes.xlsx')
client_page = client_sheet['Sheet1']

# 2. Access the CPF consultation website and use the CPF from the spreadsheet to search for the client's payment status.
driver = driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')

for line in client_page.iter_rows(min_row=2, values_only=True):
	name, value, cpf, due_date = line
	
	# xpath //tag[@atributo='valor']
	search_field = driver.find_element(By.XPATH,"//input[@id='cpfInput']")
	sleep(1)
	search_field.clear()
	search_field.send_keys(cpf)
	sleep(1)
	# 3. Check if it is up to date or overdue.
	search_button = driver.find_element(By.XPATH,"//button[@class='btn btn-custom btn-lg btn-block mt-3']")
	sleep(1)
	search_button.click()
	sleep(4)
	status = driver.find_element(By.XPATH,"//span[@id='statusLabel']")
	if status.text == 'em dia':
		# 4. If it is "up to date," get the payment date and payment method.
		payment_date = driver.find_element(By.XPATH,"//p[@id='paymentDate']")
		metodo_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentMethod']")

		payment_date_only = payment_date.text.split()[3]
		payment_method_only = metodo_pagamento.text.split()[3]

		results_sheet = openpyxl.load_workbook('planilha_fechamento.xlsx')
		results_page_sheet = results_sheet['Sheet1']

		results_page_sheet.append([name, value, cpf, due_date, 'em dia', payment_date_only, payment_method_only])

		results_sheet.save('planilha_fechamento.xlsx')
	else:
		# 5. Otherwise (if overdue), set the status as pending.
		results_sheet = openpyxl.load_workbook('planilha_fechamento.xlsx')
		results_page_sheet = results_sheet['Sheet1']

		results_page_sheet.append([name, value, cpf, due_date, 'pendente'])
		results_sheet.save('planilha_fechamento.xlsx')

